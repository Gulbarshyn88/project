from datetime import datetime
from typing import List, Dict, Union, Optional
from openpyxl import Workbook, load_workbook

ProductType = Dict[str, Union[str, float, int]]
SaleType = Dict[str, Union[List[ProductType], float, str, float, float, str]]
InventoryType = Dict[str, Union[List[ProductType], List[SaleType]]]

EXCEL_FILE = 'inventory_data.xlsx'

def save_to_excel(inventory: InventoryType, excel_file: str = EXCEL_FILE) -> None:
    wb = Workbook()
    ws = wb.active

    # Заголовки для продуктов
    ws.append(["code", "name", "price", "quantity"])
    for product in inventory['products']:
        ws.append([product['code'], product['name'], product['price'], product['quantity']])

    # Пустая строка
    ws.append([])

    # Заголовки для продаж
    ws.append(["products", "total_amount", "payment_method", "discount", "tax_rate", "date"])
    for sale in inventory['sales_history']:
        product_data = "\n".join([f"Code: {item['code']}, Name: {item['name']}, "
                                  f"Price: {item['price']}, Quantity: {item['quantity']}" for item in sale['products']])
        ws.append([product_data, sale['total_amount'], sale['payment_method'],
                   sale['discount'], sale['tax_rate'], sale['date']])

    # Сохраняем в файл
    wb.save(excel_file)
    print(f"Данные сохранены в файле: {excel_file}")


def load_from_excel(excel_file: str = EXCEL_FILE) -> InventoryType:
    try:
        wb = load_workbook(excel_file)
        ws = wb.active

        products = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                break
            product = {'code': row[0], 'name': row[1], 'price': row[2], 'quantity': row[3]}
            products.append(product)

        sales = []
        for row in ws.iter_rows(min_row=len(products) + 4, values_only=True):
            if not any(row):
                break
            product_data = [item.split(';') for item in row[0].split(',')]
            products_list = [
                {
                    'code': p[0] if p and len(p) > 0 else '',
                    'name': p[1] if len(p) > 1 else '',
                    'price': p[2] if len(p) > 2 else 0.0,
                    'quantity': p[3] if len(p) > 3 else 0
                }
                for p in product_data
            ]

            # Пробуем преобразовать дату из строки в datetime
            try:
                sale_date = datetime.strptime(row[5], "%Y-%m-%d %H:%M:%S")
            except ValueError:
                sale_date = row[5]  # Если не удается преобразовать, оставляем как строку

            sale = {'products': products_list, 'total_amount': row[1], 'payment_method': row[2],
                    'discount': row[3], 'tax_rate': row[4], 'date': sale_date}
            sales.append(sale)

        return {'products': products, 'sales_history': sales}
    except FileNotFoundError:
        return create_inventory_system()


def create_product(code: str, name: str, price: float, quantity: int) -> ProductType:
    return {'code': code, 'name': name, 'price': price, 'quantity': quantity}

def create_sale(products: List[ProductType], total_amount: float, payment_method: str,
                discount: float, tax_rate: float) -> SaleType:
    return {'products': products, 'total_amount': total_amount, 'payment_method': payment_method,
            'discount': discount, 'tax_rate': tax_rate, 'date': datetime.now().strftime("%Y-%m-%d %H:%M:%S")}

def create_inventory_system() -> InventoryType:
    return {'products': [], 'sales_history': []}

def add_product(inventory: InventoryType, code: str, name: str, price: float, quantity: int) -> None:
    new_product = create_product(code, name, price, quantity)
    inventory['products'].append(new_product)
    print(f"Товар '{name}' добавлен в систему.")

def remove_product(inventory: InventoryType, code: str) -> None:
    for product in inventory['products']:
        if product['code'] == code:
            inventory['products'].remove(product)
            print(f"Товар с кодом {code} удален из системы.")
            return
    print(f"Товар с кодом {code} не найден.")

def edit_product(inventory: InventoryType, code: str, new_name: str, new_price: float, new_quantity: int) -> None:
    for product in inventory['products']:
        if product['code'] == code:
            product['name'] = new_name
            product['price'] = new_price
            product['quantity'] = new_quantity
            print(f"Товар с кодом {code} отредактирован.")
            return
    print(f"Товар с кодом {code} не найден.")

def search_product(inventory: InventoryType, keyword: str) -> None:
    results = [product for product in inventory['products'] if keyword.lower() in product['name'].lower()]
    if not results:
        print(f"Товар с названием или кодом '{keyword}' не найден.")
    else:
        print("\nРезультаты поиска:")
        for product in results:
            print(f"Код: {product['code']}, Название: {product['name']}, Цена: {product['price']}, Количество: {product['quantity']}")

def display_inventory(inventory: InventoryType) -> None:
    print("\nТекущий инвентарь:")
    for product in inventory['products']:
        print(f"Код: {product['code']}, Название: {product['name']}, Цена: {product['price']}, Количество: {product['quantity']}")

def make_sale(inventory: InventoryType, products: List[ProductType], payment_method: str,
              discount: float = 0, tax_rate: float = 0) -> None:
    total_amount = sum(product['price'] * product['quantity'] for product in products)
    total_amount *= 1 + tax_rate / 100  # Apply tax
    total_amount *= 1 - discount / 100  # Apply discount

    for product in products:
        inventory_product = next((p for p in inventory['products'] if p['code'] == product['code']), None)
        if inventory_product is None or inventory_product['quantity'] < product['quantity']:
            print(f"Товар '{product['name']}' недоступен или недостаточно в наличии.")
            return

    for product in products:
        inventory_product = next(p for p in inventory['products'] if p['code'] == product['code'])
        inventory_product['quantity'] -= product['quantity']

    sale = create_sale(products, total_amount, payment_method, discount, tax_rate)
    inventory['sales_history'].append(sale)

    print("\nЧек:")
    for product in products:
        print(f"Название: {product['name']}, Количество: {product['quantity']}, Цена: {product['price']}")
    print(f"Общая сумма: {total_amount}, Скидка: {discount}%, Налог: {tax_rate}%, Способ оплаты: {payment_method}")
    print("Продажа добавлена в историю.")


def generate_report(inventory: InventoryType, start_date: datetime = None, 
                    end_date: datetime = None, specific_product_code: str = None) -> None:
    filtered_sales = inventory['sales_history']

    # Фильтруем продажи
    filtered_sales = [
        sale for sale in filtered_sales
        if (not start_date or datetime.strptime(sale['date'], "%Y-%m-%d %H:%M:%S") >= start_date)
        and (not end_date or datetime.strptime(sale['date'], "%Y-%m-%d %H:%M:%S") <= end_date)
        and (not specific_product_code or any(product['code'] == specific_product_code for product in sale['products']))
    ]

    # Выводим отчет
    if filtered_sales:
        print("\nОтчет:")
        for sale in filtered_sales:
            print(f"Дата: {sale['date']}, "
                  f"Общая сумма: {sale['total_amount']}, "
                  f"Скидка: {sale['discount']}%, "
                  f"Налог: {sale['tax_rate']}%, "
                  f"Способ оплаты: {sale['payment_method']}")
    else:
        print("Нет данных для отчета.")
